"""Tests for the BMG plate-reader XLSX importer and raw data writers.

The BMG reader:
    * recognises the ``Microplate End point`` sheet,
    * parses the plate grid into long-format replicas with placeholder
      concentrations ``1..N``,
    * flags the DataFrame via ``df.attrs`` so the GUI can prompt.

The measurement writers round-trip the resulting MeasurementSet through
the existing TxtReader / CsvReader.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from core.data_processing.measurement_set import MeasurementSet
from core.io import load_measurements
from core.io.formats.bmg_reader import is_bmg_workbook, parse_bmg_workbook
from core.io.formats.measurement_writer import (
    write_measurements_csv,
    write_measurements_txt,
)

_BUNDLED_BMG = Path('data/01_ZBeta_MDAP_Dopamine_pH2.xlsx')


def _write_structured_xlsx(path: Path) -> None:
    """Write a minimal non-BMG .xlsx so the dispatcher's fallback path is exercised."""
    wb = Workbook()
    ws = wb.active
    ws.append(['concentration', 'signal', 'replica'])
    for conc, sig, rep in [
        (0.0, 100.0, 0),
        (1e-6, 200.0, 0),
        (2e-6, 300.0, 0),
        (0.0, 102.0, 1),
        (1e-6, 198.0, 1),
        (2e-6, 305.0, 1),
    ]:
        ws.append([conc, sig, rep])
    wb.save(path)


class TestBMGDetection:
    def test_detects_bundled_bmg_export(self):
        if not _BUNDLED_BMG.exists():
            pytest.skip('bundled BMG fixture not available')
        wb = load_workbook(_BUNDLED_BMG, data_only=True, read_only=True)
        try:
            assert is_bmg_workbook(wb) is True
        finally:
            wb.close()

    def test_structured_xlsx_is_not_bmg(self, tmp_path):
        path = tmp_path / 'structured.xlsx'
        _write_structured_xlsx(path)
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            assert is_bmg_workbook(wb) is False
        finally:
            wb.close()


class TestBMGParse:
    def test_shape_and_placeholders(self):
        if not _BUNDLED_BMG.exists():
            pytest.skip('bundled BMG fixture not available')
        wb = load_workbook(_BUNDLED_BMG, data_only=True, read_only=True)
        try:
            df, meta = parse_bmg_workbook(wb)
        finally:
            wb.close()

        # 8 plate rows × 12 plate columns = 96 data rows
        assert set(df['replica'].unique()) == set(range(8))
        assert len(df[df['replica'] == 0]) == 12
        assert df['concentration'].min() == 1.0
        assert df['concentration'].max() == 12.0
        # Sentinel attrs so the GUI prompt fires
        assert df.attrs.get('bmg_placeholder_concentrations') is True
        assert isinstance(meta, dict)
        # Test name should be captured from the Microplate sheet metadata
        joined = ' '.join(str(v) for v in meta.values()).lower()
        assert 'zeolite' in joined or 'test_name' in {k.lower() for k in meta}

    def test_first_row_signal_values(self):
        if not _BUNDLED_BMG.exists():
            pytest.skip('bundled BMG fixture not available')
        wb = load_workbook(_BUNDLED_BMG, data_only=True, read_only=True)
        try:
            df, _meta = parse_bmg_workbook(wb)
        finally:
            wb.close()

        # Plate row A, columns 1–3: 208319, 163967, 127277 per the sheet
        rep0 = df[df['replica'] == 0].sort_values('concentration')
        assert rep0['signal'].iloc[0] == pytest.approx(208319.0)
        assert rep0['signal'].iloc[1] == pytest.approx(163967.0)
        assert rep0['signal'].iloc[2] == pytest.approx(127277.0)


class TestXlsxDispatcher:
    def test_structured_xlsx_still_works(self, tmp_path):
        """XlsxReader must fall back to the structured path for non-BMG files."""
        path = tmp_path / 'structured.xlsx'
        _write_structured_xlsx(path)
        df = load_measurements(path)
        assert set(df['replica'].unique()) == {0, 1}
        assert df.attrs.get('bmg_placeholder_concentrations') is not True

    def test_bmg_xlsx_routes_to_bmg_parser(self):
        if not _BUNDLED_BMG.exists():
            pytest.skip('bundled BMG fixture not available')
        df = load_measurements(_BUNDLED_BMG)
        assert df.attrs.get('bmg_placeholder_concentrations') is True
        assert len(df['replica'].unique()) == 8


class TestMeasurementWriters:
    def _build_ms(self) -> MeasurementSet:
        df = pd.DataFrame(
            [
                {'concentration': 0.0, 'signal': 100.0, 'replica': 0},
                {'concentration': 1e-6, 'signal': 200.0, 'replica': 0},
                {'concentration': 2e-6, 'signal': 300.0, 'replica': 0},
                {'concentration': 0.0, 'signal': 102.0, 'replica': 1},
                {'concentration': 1e-6, 'signal': 198.0, 'replica': 1},
                {'concentration': 2e-6, 'signal': 305.0, 'replica': 1},
            ]
        )
        return MeasurementSet.from_dataframe(df)

    def test_txt_round_trip(self, tmp_path):
        ms = self._build_ms()
        path = tmp_path / 'export.txt'
        write_measurements_txt(ms, path)
        reloaded = load_measurements(path)
        assert len(reloaded) == 6
        assert set(reloaded['replica'].unique()) == {0, 1}
        # Concentrations and signals survive the round trip
        rep0 = reloaded[reloaded['replica'] == 0].sort_values('concentration')
        assert rep0['concentration'].tolist() == pytest.approx([0.0, 1e-6, 2e-6])
        assert rep0['signal'].tolist() == pytest.approx([100.0, 200.0, 300.0])

    def test_csv_round_trip(self, tmp_path):
        ms = self._build_ms()
        path = tmp_path / 'export.csv'
        write_measurements_csv(ms, path)
        reloaded = load_measurements(path)
        assert len(reloaded) == 6
        assert set(reloaded['replica'].unique()) == {0, 1}
        rep1 = reloaded[reloaded['replica'] == 1].sort_values('concentration')
        assert rep1['signal'].tolist() == pytest.approx([102.0, 198.0, 305.0])


class TestBMGRoundTripAfterConcentrationFix:
    """Load BMG, replace placeholder concentrations, export, reload."""

    def test_full_round_trip(self, tmp_path):
        if not _BUNDLED_BMG.exists():
            pytest.skip('bundled BMG fixture not available')
        df = load_measurements(_BUNDLED_BMG)
        ms = MeasurementSet.from_dataframe(
            df,
            source_file=str(_BUNDLED_BMG),
            bmg_placeholder_concentrations=True,
        )
        assert ms.metadata.get('bmg_placeholder_concentrations') is True

        # Rebuild with real concentrations (1 nM–12 nM as an example)
        real_conc = np.array([i * 1e-9 for i in range(1, ms.n_points + 1)])
        rows = []
        for rep_idx, (_rid, signal) in enumerate(ms.iter_replicas(active_only=False)):
            for c, s in zip(real_conc, signal):
                rows.append({'concentration': c, 'signal': float(s), 'replica': rep_idx})
        fixed = MeasurementSet.from_dataframe(pd.DataFrame(rows))

        path = tmp_path / 'bmg_roundtrip.txt'
        write_measurements_txt(fixed, path)
        reloaded_df = load_measurements(path)
        reloaded = MeasurementSet.from_dataframe(reloaded_df)

        assert reloaded.n_replicas == fixed.n_replicas
        assert reloaded.n_points == fixed.n_points
        np.testing.assert_allclose(reloaded.concentrations, fixed.concentrations)
        np.testing.assert_allclose(reloaded.signals, fixed.signals)
