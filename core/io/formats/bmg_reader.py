"""BMG plate reader export parser.

BMG CLARIOstar / FLUOstar / PHERAstar instruments export fluorescence
measurements as ``.xlsx`` workbooks with two sheets:

* ``Microplate End point`` — the plate grid, with a few metadata rows at
  the top, a "Raw Data (...)" header, a column-number header row
  (``1, 2, ..., N``) and then one row per well letter (``A..H`` for a
  96-well plate, ``A..P`` for a 384-well plate).
* ``Protocol Information`` — free-form metadata.

BMG exports do **not** contain concentration information: each plate
row is one replica of the titration, each plate column is one titration
point, and the mapping from column index to concentration is kept
outside the file. This reader assigns placeholder concentrations
``1..N`` and flags the resulting DataFrame via ``df.attrs`` so the GUI
can prompt the user to supply real values.
"""

from __future__ import annotations

import logging
import re
import string
from pathlib import Path
from typing import Any, Dict, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

logger = logging.getLogger(__name__)

#: Metadata key used on ``MeasurementSet.metadata`` and ``df.attrs`` to flag
#: datasets whose concentrations are the placeholder plate-column positions
#: ``1..N`` rather than real titration values. The GUI uses this flag to
#: prompt for real concentrations and to refuse to run a fit.
BMG_PLACEHOLDER_KEY = 'bmg_placeholder_concentrations'
BMG_METADATA_KEY = 'bmg_metadata'

_PLATE_ROW_LETTERS = set(string.ascii_uppercase[:16])  # A..P, covers 96 and 384 well
_METADATA_PREFIXES = (
    'test id',
    'test name',
    'date',
    'time',
    'user',
    'path',
    'fluorescence',
    'measurement type',
)


def is_bmg_workbook(wb: Workbook) -> bool:
    """Return True if *wb* looks like a BMG plate-reader export.

    Detection is cheap: find a sheet whose name starts with
    ``"microplate end"`` (case-insensitive) and scan its first ~30 rows
    for a column-number header row (col A empty, cols B..M holding the
    integers 1, 2, 3, ...).
    """
    target = _find_bmg_sheet(wb)
    if target is None:
        return False
    return _find_header_row(target) is not None


def parse_bmg_workbook(wb: Workbook) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Parse a BMG workbook into a long-format DataFrame plus metadata.

    Returns
    -------
    df : pd.DataFrame
        Columns ``concentration``, ``signal``, ``replica``. Concentrations
        are bare integers ``1..N`` (placeholders — real values must be
        supplied by the user).
    metadata : dict
        Free-form metadata extracted from the sheet headers and, if
        present, the ``Protocol Information`` sheet.
    """
    sheet = _find_bmg_sheet(wb)
    if sheet is None:
        raise ValueError('Workbook does not contain a Microplate End point sheet.')

    header_row = _find_header_row(sheet)
    if header_row is None:
        raise ValueError('Could not locate the column-number header row in BMG sheet.')

    # Column headers: col B.. onwards hold 1, 2, 3, ...
    header_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    column_indices: list[int] = []
    for j, value in enumerate(header_cells[1:], start=2):  # start=2 → col B
        if isinstance(value, (int, float)) and float(value).is_integer() and int(value) == len(column_indices) + 1:
            column_indices.append(j)
        else:
            break
    n_cols = len(column_indices)
    if n_cols == 0:
        raise ValueError('BMG header row contained no column numbers.')

    # Data rows: col A holds a single plate-row letter, cols B..(B+n_cols) hold signals
    replicas: list[list[float]] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        label = row[0]
        if not isinstance(label, str) or label.strip().upper() not in _PLATE_ROW_LETTERS:
            break
        signals: list[float] = []
        for col_idx in column_indices:
            value = row[col_idx - 1]
            if value is None:
                signals.append(float('nan'))
            else:
                signals.append(float(value))
        replicas.append(signals)

    if not replicas:
        raise ValueError('BMG sheet had a header row but no plate-row data rows.')

    # Build long-format DataFrame with placeholder concentrations 1..N
    rows = []
    for replica_idx, signals in enumerate(replicas):
        for col_idx, signal in enumerate(signals, start=1):
            if pd.isna(signal):
                continue
            rows.append(
                {
                    'concentration': float(col_idx),
                    'signal': signal,
                    'replica': replica_idx,
                }
            )
    df = pd.DataFrame(rows)

    # Annotate the DataFrame so downstream loaders can detect BMG data
    df.attrs[BMG_PLACEHOLDER_KEY] = True
    df.attrs[BMG_METADATA_KEY] = _extract_metadata(wb, sheet)
    return df, df.attrs[BMG_METADATA_KEY]


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------


def _find_bmg_sheet(wb: Workbook):
    for name in wb.sheetnames:
        if name.strip().lower().startswith('microplate end'):
            return wb[name]
    return None


def _find_header_row(sheet) -> int | None:
    """Return the 1-based row index of the column-number header, or None."""
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        if not row or len(row) < 2:
            continue
        if row[0] is not None:
            continue
        # Expect 1, 2, 3, ... in cols B.. — require at least 4 consecutive ints
        seq = []
        for v in row[1:]:
            if isinstance(v, (int, float)) and float(v).is_integer() and int(v) == len(seq) + 1:
                seq.append(int(v))
            else:
                break
        if len(seq) >= 4:
            return row_idx
    return None


def _extract_metadata(wb: Workbook, sheet) -> Dict[str, Any]:
    """Pull a small dict of free-form metadata from the BMG workbook."""
    meta: Dict[str, Any] = {}

    # Scan the first 12 rows of col A for 'Key: Value' style strings
    for row in sheet.iter_rows(min_row=1, max_row=12, max_col=1, values_only=True):
        cell = row[0]
        if not isinstance(cell, str):
            continue
        for prefix in _METADATA_PREFIXES:
            if cell.strip().lower().startswith(prefix):
                key, _, value = cell.partition(':')
                meta[key.strip().lower().replace(' ', '_')] = value.strip() or key.strip()
                break

    # Also grab a few useful rows from the Protocol Information sheet
    for name in wb.sheetnames:
        if name.strip().lower().startswith('protocol'):
            proto = wb[name]
            for row in proto.iter_rows(min_row=1, max_row=40, values_only=True):
                if len(row) < 2:
                    continue
                key, value = row[0], row[1]
                if isinstance(key, str) and isinstance(value, (str, int, float)):
                    norm_key = re.sub(r'[^a-z0-9]+', '_', key.strip().lower()).strip('_')
                    if norm_key and norm_key not in meta:
                        meta[norm_key] = value
            break

    return meta
